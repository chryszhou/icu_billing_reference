"""Regenerate data.js from the workbook. Run from the repo root:  python3 gen_data.py"""
import json
from openpyxl import load_workbook

wb = load_workbook('downloads/ICU_Diagnosis_DRG_Reference_2026.xlsx', data_only=False)

def val(ws,r,c):
    v=ws.cell(row=r,column=c).value
    return v if v not in ("",None) else None

ws=wb['Diagnoses by Organ System']
organ=[]; cur=None
for r in range(4, ws.max_row+1):
    a,b,c,d=[val(ws,r,i) for i in (1,2,3,4)]
    if a and not b:
        if a=='ICD-10': continue
        cur={"section":a,"rows":[]}; organ.append(cur); continue
    if a and b and a!='ICD-10' and cur is not None:
        cur["rows"].append({"code":a,"name":b,"status":c or "","tip":d or ""})

ws=wb['DRG Value Ranking']
base=val(ws,4,4) or 6752.61
drg=[]
for r in range(6, ws.max_row+1):
    a,b,c,dd=[val(ws,r,i) for i in (1,2,3,4)]
    if a=='Condition family': continue
    if b and isinstance(dd,(int,float)):
        drg.append({"family":a,"drg":str(b),"tier":c or "","weight":float(dd)})

def simple(sheet, cols, header0):
    ws=wb[sheet]; out=[]
    for r in range(4, ws.max_row+1):
        vals=[val(ws,r,i) for i in (1,2,3,4)]
        if vals[0]==header0: continue
        if vals[1] is None: continue
        out.append({cols[i]:(vals[i] or "") for i in range(4)})
    return out

data={"base":base,"organ":organ,"drg":drg,
      "complex":simple('Complex Care & Long LOS',["code","name","status","tip"],"Code"),
      "cirrhosis":simple('Cirrhosis Decompensations',["name","code","status","tip"],"Decompensation"),
      "malnutrition":simple('Malnutrition',["name","code","status","tip"],"Severity"),
      "trachvent":simple('Chronic Trach & Vent',["name","code","status","tip"],"Item")}

with open('data.js','w') as f:
    f.write("window.SITE = "+json.dumps(data,ensure_ascii=False,indent=1)+";\n")
print("data.js regenerated")
